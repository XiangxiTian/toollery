from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable


DEFAULT_SHEET = "Sheet1"
DEFAULT_CATEGORY = "vehicle_status_query"


@dataclass(frozen=True)
class MethodParameter:
    type_name: str
    name: str


@dataclass(frozen=True)
class GetterSignature:
    return_type: str
    name: str
    parameters: list[MethodParameter]


def main() -> None:
    args = parse_args()
    rows = load_xlsx_rows(
        args.input,
        sheet_name=args.sheet,
        header_row=args.header_row,
        data_start_row=args.data_start_row,
    )
    tools = build_status_tools(
        rows,
        category=args.category,
        include_without_getter=not args.query_only,
    )
    write_tools_json(args.out, tools)
    print(
        json.dumps(
            {
                "input": str(Path(args.input).resolve()),
                "out": str(Path(args.out).resolve()),
                "sheet": args.sheet,
                "tools": len(tools),
                "parameters": sum(len(tool["parameters"]["properties"]) for tool in tools),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert vehicle status-query signals to BFCL-style tool definitions.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--input", required=True, help="Input .xlsx workbook.")
    parser.add_argument("--out", required=True, help="Output JSON file with a top-level tools array.")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Worksheet name.")
    parser.add_argument("--header-row", type=int, default=1, help="1-based row containing column headers.")
    parser.add_argument("--data-start-row", type=int, default=4, help="1-based first data row.")
    parser.add_argument(
        "--category",
        default=DEFAULT_CATEGORY,
        help="Optional category value attached to every converted tool. Use empty string to omit.",
    )
    parser.add_argument(
        "--query-only",
        action="store_true",
        help="Only emit rows that have a get...(...) method in the northbound API cell.",
    )
    return parser.parse_args()


def load_xlsx_rows(
    path: str | Path,
    sheet_name: str = DEFAULT_SHEET,
    header_row: int = 1,
    data_start_row: int = 4,
) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            "Reading .xlsx files requires openpyxl. Run this script with the Codex workspace Python "
            "runtime or install openpyxl in your Python environment."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}. Available sheets: {', '.join(workbook.sheetnames)}")
    sheet = workbook[sheet_name]
    headers = _unique_headers([_clean_text(cell.value) for cell in sheet[header_row]])
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=data_start_row, values_only=True):
        if not any(_clean_text(value) for value in values):
            continue
        rows.append({headers[index]: value for index, value in enumerate(values) if index < len(headers)})
    return rows


def build_status_tools(
    rows: Iterable[dict[str, Any]],
    category: str | None = DEFAULT_CATEGORY,
    include_without_getter: bool = True,
) -> list[dict[str, Any]]:
    grouped_rows: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        name = _first_text(row, "vspec信号")
        if name:
            grouped_rows.setdefault(name, []).append(row)

    tools: list[dict[str, Any]] = []
    for name, same_signal_rows in grouped_rows.items():
        row_getters = [
            (row, parse_getter_signatures(_first_text(row, "北向API")))
            for row in same_signal_rows
        ]
        if not include_without_getter:
            row_getters = [(row, getters) for row, getters in row_getters if getters]
        if not row_getters:
            continue
        getters = [getter for _, getter_list in row_getters for getter in getter_list]
        tool = {
            "name": name,
            "description": _merged_description(row_getters),
            "parameters": _parameters(row_getters),
        }
        if category:
            tool["category"] = category
        tools.append(tool)
    return tools


def parse_getter_signatures(api_text: str) -> list[GetterSignature]:
    pattern = re.compile(
        r"\b(?P<return>[A-Za-z_][\w.<>\[\]?]*)\s+"
        r"(?P<name>get[A-Za-z_]\w*)\s*"
        r"\((?P<params>[^)]*)\)"
    )
    signatures: list[GetterSignature] = []
    for match in pattern.finditer(api_text or ""):
        signatures.append(
            GetterSignature(
                return_type=match.group("return"),
                name=match.group("name"),
                parameters=_parse_parameters(match.group("params")),
            )
        )
    return signatures


def write_tools_json(path: str | Path, tools: list[dict[str, Any]]) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps({"tools": tools}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _parse_parameters(params_text: str) -> list[MethodParameter]:
    params: list[MethodParameter] = []
    for raw_param in (params_text or "").split(","):
        raw_param = re.sub(r"@\w+(?:\([^)]*\))?\s*", "", raw_param).strip()
        if not raw_param:
            continue
        parts = raw_param.split()
        if len(parts) < 2:
            continue
        type_name = " ".join(parts[:-1]).strip()
        name = parts[-1].strip()
        name = name.lstrip("...").strip()
        if type_name and name:
            params.append(MethodParameter(type_name=type_name, name=name))
    return params


def _parameters(row_getters: list[tuple[dict[str, Any], list[GetterSignature]]]) -> dict[str, Any]:
    properties: dict[str, dict[str, str]] = {}
    for row, getters in row_getters:
        for getter in getters:
            for param in getter.parameters:
                properties.setdefault(
                    param.name,
                    {
                        "type": _bfcl_type(param.type_name),
                        "description": _parameter_description(row, param),
                    },
                )
    return {"type": "dict", "properties": properties, "required": []}


def _merged_description(row_getters: list[tuple[dict[str, Any], list[GetterSignature]]]) -> str:
    descriptions: list[str] = []
    seen: set[str] = set()
    for row, getters in row_getters:
        description = _description(row, getters)
        if description and description not in seen:
            descriptions.append(description)
            seen.add(description)
    return " ".join(descriptions)


def _description(row: dict[str, Any], getters: list[GetterSignature]) -> str:
    module = _first_text(row, "所属模块")
    signal = _first_text(row, "信号名称")
    value_desc = _first_text(row, "北向API接口枚举值或范围")
    vspec_type = _first_text(row, "vspec类型")
    propid = _first_text(row, "propid")
    api_methods = ", ".join(getter.name for getter in getters)
    parts = []
    if module and signal:
        parts.append(f"查询{module}模块的{signal}。")
    elif signal:
        parts.append(f"查询{signal}。")
    if value_desc:
        parts.append(f"返回值或范围：{_trim_sentence(_single_line(value_desc))}。")
    if vspec_type:
        parts.append(f"VSpec类型：{vspec_type}。")
    if propid:
        parts.append(f"PropID：{propid}。")
    if api_methods:
        parts.append(f"北向查询API：{api_methods}。")
    return "".join(parts)


def _parameter_description(row: dict[str, Any], param: MethodParameter) -> str:
    signal = _first_text(row, "信号名称") or "该状态"
    return f"{param.type_name} {param.name}，用于指定查询{signal}时的车辆区域或位置。"


def _unique_headers(headers: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    out: list[str] = []
    for index, header in enumerate(headers, start=1):
        name = header or f"col_{index}"
        counts[name] = counts.get(name, 0) + 1
        out.append(name if counts[name] == 1 else f"{name}_{counts[name]}")
    return out


def _first_text(row: dict[str, Any], *keys: str) -> str:
    for key in keys:
        text = _clean_text(row.get(key))
        if text:
            return text
    return ""


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"none", "nan"} else text


def _single_line(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def _trim_sentence(value: str) -> str:
    return value.rstrip("。.;；")


def _bfcl_type(value: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "", value.lower())
    if normalized in {"int", "integer", "long", "short", "byte", "integerzone"}:
        return "integer"
    if normalized in {"float", "double", "number", "decimal"}:
        return "number"
    if normalized in {"bool", "boolean"}:
        return "boolean"
    if "list" in normalized or "array" in normalized:
        return "array"
    if normalized in {"dict", "object", "map"}:
        return "dict"
    return "string"


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        raise
