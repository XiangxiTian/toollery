from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any, Iterable


DEFAULT_SHEET = "main_data"
DEFAULT_CATEGORY = "vehicle_control"


def main() -> None:
    args = parse_args()
    rows = load_xlsx_rows(
        args.input,
        sheet_name=args.sheet,
        header_row=args.header_row,
        data_start_row=args.data_start_row,
    )
    tools = build_car_tools(rows, category=args.category)
    write_tools_json(args.out, tools)
    print(
        json.dumps(
            {
                "input": str(Path(args.input).resolve()),
                "out": str(Path(args.out).resolve()),
                "sheet": args.sheet,
                "tools": len(tools),
                "parameters": sum(len(tool["parameters"]["properties"]) for tool in tools),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert the vehicle-control tool workbook to BFCL-style tool definitions.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--input", required=True, help="Input .xlsx workbook.")
    parser.add_argument("--out", required=True, help="Output JSON file with a top-level tools array.")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Worksheet name.")
    parser.add_argument("--header-row", type=int, default=1, help="1-based row containing column headers.")
    parser.add_argument("--data-start-row", type=int, default=3, help="1-based first data row.")
    parser.add_argument(
        "--category",
        default=DEFAULT_CATEGORY,
        help="Optional category value attached to every converted tool. Use empty string to omit.",
    )
    return parser.parse_args()


def load_xlsx_rows(
    path: str | Path,
    sheet_name: str = DEFAULT_SHEET,
    header_row: int = 1,
    data_start_row: int = 3,
) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            "Reading .xlsx files requires openpyxl. Run this script with the Codex workspace Python "
            "runtime or install openpyxl in your Python environment."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}. Available sheets: {', '.join(workbook.sheetnames)}")
    sheet = workbook[sheet_name]
    header_values = [
        _clean_text(cell.value) or f"col_{index}"
        for index, cell in enumerate(sheet[header_row], start=1)
    ]
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=data_start_row, values_only=True):
        if not any(_clean_text(value) for value in values):
            continue
        row = {header_values[index]: value for index, value in enumerate(values) if index < len(header_values)}
        rows.append(row)
    return rows


def build_car_tools(rows: Iterable[dict[str, Any]], category: str | None = DEFAULT_CATEGORY) -> list[dict[str, Any]]:
    tools: list[dict[str, Any]] = []
    current_tool: dict[str, Any] | None = None

    for row in rows:
        tool_name = _first_text(row, "actionNameEn", "actionNameZh")
        tool_description = _first_text(row, "actionDescribe")
        if tool_name:
            current_tool = _new_tool(tool_name, tool_description, category)
            tools.append(current_tool)
        if current_tool is None:
            continue

        param_name = _first_text(row, "inputNameEn", "inputNameZh")
        if not param_name:
            continue
        properties = current_tool["parameters"]["properties"]
        properties.setdefault(
            param_name,
            {
                "type": _bfcl_type(_first_text(row, "dataType", "inputType")),
                "description": _first_text(row, "inputDescribe"),
            },
        )

    return tools


def write_tools_json(path: str | Path, tools: list[dict[str, Any]]) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps({"tools": tools}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def load_bfcl_tools_file(path: str | Path) -> list[dict[str, Any]]:
    source = Path(path)
    text = source.read_text(encoding="utf-8").strip()
    if not text:
        return []
    if source.suffix == ".jsonl":
        return [json.loads(line) for line in text.splitlines() if line.strip()]
    data = json.loads(text)
    if isinstance(data, dict) and isinstance(data.get("tools"), list):
        return [_normalize_tool(item) for item in data["tools"] if isinstance(item, dict)]
    if isinstance(data, list):
        return [_normalize_tool(item) for item in data if isinstance(item, dict)]
    raise ValueError(f"Unsupported tools file shape in {source}; expected JSON tools array or JSONL tool rows.")


def _new_tool(name: str, description: str, category: str | None) -> dict[str, Any]:
    tool = {
        "name": name,
        "description": description,
        "parameters": {"type": "dict", "properties": {}, "required": []},
    }
    if category:
        tool["category"] = category
    return tool


def _normalize_tool(tool: dict[str, Any]) -> dict[str, Any]:
    out = {
        "name": str(tool["name"]),
        "description": str(tool.get("description", "")),
        "parameters": dict(tool.get("parameters", {})),
    }
    if "category" in tool:
        out["category"] = tool["category"]
    return out


def _first_text(row: dict[str, Any], *keys: str) -> str:
    for key in keys:
        text = _clean_text(row.get(key))
        if text:
            return text
    return ""


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"none", "nan"} else text


def _bfcl_type(value: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "", value.lower())
    if normalized in {"int", "integer", "long"}:
        return "integer"
    if normalized in {"float", "double", "number", "decimal"}:
        return "number"
    if normalized in {"bool", "boolean"}:
        return "boolean"
    if normalized in {"array", "list"}:
        return "array"
    if normalized in {"dict", "object", "map"}:
        return "dict"
    return "string"


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        raise
